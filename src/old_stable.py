import tkinter as tk
from tkinter import filedialog, messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.drawing.image import Image
from copy import copy
import unicodedata, sys, os, subprocess

base_file = None
extrato_file = None
report_files = []

thick = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )

no_border = Border(
            left=Side(style=None),
            right=Side(style=None),
            top=Side(style=None),
            bottom=Side(style=None)
        )

border = Border(
            left=Side(style='hair'),
            right=Side(style='hair'),
            top=Side(style='hair'),
            bottom=Side(style='hair')
        )

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

alinhamentos = {
    "A": "right",    # Nº
    "B": "left",     # RESUMO DO GASTO
    "C": "left",     # RUBRICA
    "D": "left",     # FORNECEDOR
    "E": "center",   # PC
    "F": "center",   # CNPJ ou CPF
    "G": "center",   # DATA DA NF
    "H": "center",   # (número) NF/RECIBO
    "I": "center",   # Nº EXTRATO
    "J": "center",   # DATA DO PAGAMENTO
    "K": "right",    # VALOR PAGO
    "L": "center",   # CÓD. PEDIDO
    "M": "center",   # CÓD. RMS
    "N": "center",   # Mês_pagto
    "O": "left",     # HISTÓRICO
    "P": "left"      # EXCEÇÕES DO ALGORITMO
}

# ==============================
# DICIONÁRIO DE COLUNAS
# ==============================

def normalizar_texto(texto):
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("ASCII")
    texto = texto.replace("º", "o")
    return texto

associacoes = {
    "no": "Nº",
    "resumo do gasto": "RESUMO DO GASTO",
    "rubrica": "RUBRICA",
    "fornecedor": "FORNECEDOR",
    "pc": "PC",
    "cnpj ou cpf": "CNPJ ou CPF",
    "data da nf": "DATA DA NF",
    "(numero) nf/recibo": "(número) NF/RECIBO",
    "no extrato": "Nº EXTRATO",
    "data do pagamento": "DATA DO PAGAMENTO",
    "valor pago": "VALOR PAGO",
    "cod. pedido": "CÓD. PEDIDO",
    "cod. rms": "CÓD. RMS",
    "mes_pagto": "Mês_pagto",
    "historico": "HISTÓRICO",
    "excecoes do algoritmo": "EXCEÇÕES DO ALGORITMO"
}

COLUNAS_PADRAO = [
    "Nº",
    "RESUMO DO GASTO",
    "RUBRICA",
    "FORNECEDOR",
    "PC",
    "CNPJ ou CPF",
    "DATA DA NF",
    "(número) NF/RECIBO",
    "Nº EXTRATO",
    "DATA DO PAGAMENTO",
    "VALOR PAGO",
    "CÓD. PEDIDO",
    "CÓD. RMS",
    "Mês_pagto",
    "HISTÓRICO",
    "EXCEÇÕES DO ALGORITMO"
]

# ==============================
# FUNÇÕES AUXILIARES
# ==============================

def encontrar_coluna(df, nomes_possiveis, obrigatoria=True):
    """
    Procura múltiplos nomes de coluna no DataFrame
    e retorna a primeira encontrada.
    """

    colunas = [normalizar_texto(c) for c in df.columns]

    for nome in nomes_possiveis:
        nome_norm = normalizar_texto(nome)

        if nome_norm in colunas:
            return nome_norm

    if obrigatoria:
        raise Exception(
            f"Nenhuma das colunas foi encontrada:\n{nomes_possiveis}"
        )

    return None

def ler_excel(path, **kwargs):
    ext = os.path.splitext(path)[1].lower()

    if ext == ".xls":
        return pd.read_excel(
            path,
            engine="xlrd",
            **kwargs
        )

    return pd.read_excel(
        path,
        engine="openpyxl",
        **kwargs
    )

def detectar_header_extrato(path):

    df_temp = ler_excel(
        path,
        header=None
    )

    for i in range(min(20, len(df_temp))):

        linha = [
            normalizar_texto(v)
            for v in df_temp.iloc[i].tolist()
        ]

        # Tipo A
        if (
            "numero documento" in linha and
            "valor r$" in linha
        ):
            return i

        # Tipo B
        if (
            "doc" in linha and
            "valor" in linha
        ):
            return i

    raise Exception(
        "Não foi possível localizar o cabeçalho do extrato."
    )

def validar_planilha_base(ws):

    linha_header = None

    for row in range(1, 11):

        valor = ws.cell(row=row, column=1).value

        if normalizar_texto(valor) == "no":
            linha_header = row
            break

    if linha_header is None:
        raise Exception(
            'Cabeçalho não encontrado.'
        )

    erros = []

    for idx, nome in enumerate(COLUNAS_PADRAO, start=1):

        valor = ws.cell(
            row=linha_header,
            column=idx
        ).value

        if normalizar_texto(valor) != normalizar_texto(nome):

            erros.append(
                f'Coluna {idx}: esperado "{nome}"'
            )

    if erros:

        raise Exception(
            "Planilha base inválida:\n\n" +
            "\n".join(erros)
        )

    return linha_header

def obter_info_tabela(ws, linha_header):

    mapa = {}

    for col in range(1, ws.max_column + 1):

        valor = ws.cell(
            row=linha_header,
            column=col
        ).value

        if valor in [None, ""]:
            continue

        nome = normalizar_texto(valor)

        if nome not in mapa:
            mapa[nome] = []

        mapa[nome].append(col)

    ultima_linha = detectar_ultima_linha(
        ws,
        mapa,
        linha_header + 1
    )

    ultima_coluna = ws.max_column

    col_letter = get_column_letter(
        ultima_coluna
    )

    matriz_ref = (
        f"'{ws.title}'!"
        f"$A${linha_header}:"
        f"${col_letter}${ultima_linha}"
    )

    return {
        "linha_header": linha_header,
        "mapa": mapa,
        "ultima_linha": ultima_linha,
        "ultima_coluna": ultima_coluna,
        "col_letter": col_letter,
        "matriz_ref": matriz_ref
    }

def indice_coluna_vlookup(mapa, nome_coluna):

    nome_normalizado = normalizar_texto(
        nome_coluna
    )

    if nome_normalizado not in mapa:

        raise Exception(
            f'Coluna "{nome_coluna}" não encontrada na planilha.'
        )

    return mapa[nome_normalizado][0]

def criar_folha_de_rosto(wb, ws, linha_header):

    info = obter_info_tabela(ws, linha_header)
    mapa = info["mapa"]
    
    # =========================================================
    # MATRIZ DINÂMICA
    # =========================================================
    #
    # Em vez de travar até a última linha atual,
    # deixa a matriz ir até o final da coluna.
    #
    # Isso evita #N/D após emendas de planilhas.
    #
    # Exemplo:
    # 'Planilha'!$A:$R
    #
    # =========================================================

    ultima_coluna = info["ultima_coluna"]
    col_letter = get_column_letter(ultima_coluna)
    ultima_linha = detectar_ultima_linha(ws, mapa, linha_header + 1)

    MATRIZ_REF = (
        f"'{ws.title}'!"
        f"$A${linha_header}:"
        f"${col_letter}${ultima_linha}"
    )

    capa = wb.create_sheet(title="Folha de Rosto", index=1)
    capa.sheet_view.showGridLines = False

    for row in capa.iter_rows():
        for cell in row:
            cell.border = no_border
            cell.alignment = Alignment(wrap_text=True)

    caminho_header = resource_path("header.png")

    if os.path.exists(caminho_header):
        img_h = Image(caminho_header)
        img_h.width = cm_para_px(25.4)
        img_h.height = cm_para_px(6.91)
        capa.add_image(img_h, "A1")

    else:
        messagebox.showwarning(
            "Imagem não encontrada",
            "O cabeçalho da folha de rosto não foi localizado."
        )

    caminho_bottom = resource_path("bottom.png")

    if os.path.exists(caminho_bottom):
        img_b = Image(caminho_bottom)
        img_b.width = cm_para_px(25.35)
        img_b.height = cm_para_px(5)
        capa.add_image(img_b, "A46")

    else:
        messagebox.showwarning(
            "Imagem não encontrada",
            "O rodapé da folha de rosto não foi localizado."
        )

    bold = Font(bold=True)

    # =========================================================
    # CAMPOS FIXOS
    # =========================================================

    capa["Q6"] = "LOC"

    # Primeira linha de dados
    capa["P6"] = linha_header + 1

    capa["D21"] = "Instituição Executora:"
    capa["D22"] = "CNPJ:"
    capa["D23"] = "Termo de Parceria:"
    capa["D24"] = "Projeto:"
    capa["D25"] = "PC:"

    capa["G21"] = "Instituto Hardware Br - HBR"
    capa["G22"] = "09.429.074/0001-12"
    capa["G25"] = entry_pc.get()

    # =========================================================
    # TABELA
    # =========================================================

    capa["C32"] = "Natureza do Dispêndio"
    capa["C33"] = "Favorecido"
    capa["H33"] = "CNPJ OU CPF"
    capa["J33"] = "Nº Extrato"
    capa["C35"] = "NF/ND"
    capa["E35"] = "Data de emissão da NF/ND"
    capa["H35"] = "Data do pagamento"
    capa["J35"] = "Valor"

    capa["C32"].font = bold

    # =========================================================
    # FÓRMULAS
    # =========================================================
    #
    # FALSE = correspondência exata
    #
    # Sem isso o VLOOKUP pode retornar #N/D
    # ou buscar linhas erradas.
    #
    # =========================================================

    capa["Q12"] = (
        f'=VLOOKUP(P6,{MATRIZ_REF},'
        f'{indice_coluna_vlookup(mapa, "RESUMO DO GASTO")},FALSE)'
    )
    capa["F32"] = (
        f'=VLOOKUP(P6,{MATRIZ_REF},'
        f'{indice_coluna_vlookup(mapa, "RUBRICA")},FALSE)'
    )
    capa["C34"] = (
        f'=VLOOKUP(P6,{MATRIZ_REF},'
        f'{indice_coluna_vlookup(mapa, "FORNECEDOR")},FALSE)'
    )
    capa["H34"] = (
        f'=VLOOKUP(P6,{MATRIZ_REF},'
        f'{indice_coluna_vlookup(mapa, "CNPJ OU CPF")},FALSE)'
    )
    capa["J34"] = (
        f'=VLOOKUP(P6,{MATRIZ_REF},'
        f'{indice_coluna_vlookup(mapa, "Nº EXTRATO")},FALSE)'
    )
    capa["J34"].number_format = '#,##0'
    capa["C36"] = (
        f'=VLOOKUP(P6,{MATRIZ_REF},'
        f'{indice_coluna_vlookup(mapa, "(número) NF/RECIBO")},FALSE)'
    )
    capa["E36"] = (
        f'=VLOOKUP(P6,{MATRIZ_REF},'
        f'{indice_coluna_vlookup(mapa, "DATA DA NF")})'
    )
    capa["H36"] = (
        f'=VLOOKUP(P6,{MATRIZ_REF},'
        f'{indice_coluna_vlookup(mapa, "DATA DO PAGAMENTO")})'
    )
    capa["J36"] = (
        f'=VLOOKUP(P6,{MATRIZ_REF},'
        f'{indice_coluna_vlookup(mapa, "VALOR PAGO")},FALSE)'
    )
    capa["J36"].number_format = 'R$ #,##0.00'

    # =========================================================
    # MERGES
    # =========================================================

    def merge(ws, cells):
        ws.merge_cells(cells)

    merge(capa, "D21:F21")
    merge(capa, "G21:I21")

    merge(capa, "D22:F22")
    merge(capa, "G22:I22")

    merge(capa, "D23:F23")
    merge(capa, "G23:I23")

    merge(capa, "D24:F24")
    merge(capa, "G24:L24")

    merge(capa, "D25:F25")
    merge(capa, "G25:I25")

    merge(capa, "C32:E32")
    merge(capa, "F32:L32")

    merge(capa, "C33:G33")
    merge(capa, "H33:I33")
    merge(capa, "J33:L33")

    merge(capa, "C34:G34")
    merge(capa, "H34:I34")
    merge(capa, "J34:L34")

    merge(capa, "C35:D35")
    merge(capa, "E35:G35")
    merge(capa, "H35:I35")
    merge(capa, "J35:L35")

    merge(capa, "C36:D36")
    merge(capa, "E36:G36")
    merge(capa, "H36:I36")
    merge(capa, "J36:L36")

    # =========================================================
    # ESTILOS
    # =========================================================

    for c in ["D21", "D22", "D23", "D25"]:
        capa[c].font = bold
        capa[c].alignment = Alignment(
            horizontal="right",
            wrap_text=True
        )

    style_range(capa, "C32:E32", thick, gray, center)
    style_range(capa, "F32:L32", thick, None, center)

    style_range(capa, "C33:G33", thick, gray, center)
    style_range(capa, "H33:I33", thick, gray, center)
    style_range(capa, "J33:L33", thick, gray, center)

    style_range(capa, "C34:G34", thick, None, center)
    style_range(capa, "H34:I34", thick, None, center)
    style_range(capa, "J34:L34", thick, None, center)

    style_range(capa, "C35:D35", thick, gray, center)
    style_range(capa, "E35:G35", thick, gray, center)
    style_range(capa, "H35:I35", thick, gray, center)
    style_range(capa, "J35:L35", thick, gray, center)

    style_range(capa, "C36:D36", thick, None, center)
    style_range(capa, "E36:G36", thick, None, center)
    style_range(capa, "H36:I36", thick, None, center)
    style_range(capa, "J36:L36", thick, None, center)

    for r in range(21, 26):

        capa.cell(
            row=r,
            column=6
        ).border = Border(
            right=Side(style='thick')
        )

        capa.cell(
            row=r,
            column=7
        ).border = Border(
            left=Side(style='thick')
        )

    capa.row_dimensions[24].height = 42
    capa.row_dimensions[32].height = 46.5
    capa.row_dimensions[34].height = 31.2

    capa["D24"].alignment = Alignment(horizontal="right", vertical="top")
    capa["D24"].font = bold
    capa["G24"].alignment = Alignment(vertical="top")

    capa["F32"].alignment = Alignment(
        wrap_text=True,
        horizontal="center",
        vertical="center"
    )

    capa["C34"].alignment = Alignment(
        wrap_text=True,
        horizontal="center",
        vertical="center"
    )

    for cell in ["E36", "H36"]:
        capa[cell].number_format = 'dd/mm/yyyy'

def detectar_ultima_linha(ws, mapa, linha_inicio):

    col_valor_pago = mapa.get(
        "valor pago",
        [None]
    )[0]

    if not col_valor_pago:
        return linha_inicio - 1

    ultima_linha = linha_inicio - 1

    row = linha_inicio

    while row <= ws.max_row:

        valor = ws.cell(
            row=row,
            column=col_valor_pago
        ).value

        if valor not in [None, ""]:

            ultima_linha = row

        else:

            vazias = 0

            for r in range(
                row,
                min(row + 5, ws.max_row + 1)
            ):

                v = ws.cell(
                    row=r,
                    column=col_valor_pago
                ).value

                if v in [None, ""]:
                    vazias += 1

            if vazias >= 5:
                break

        row += 1

    return ultima_linha

def copiar_estilo(origem, destino):

    if origem.has_style:
        destino.font = copy(origem.font)
        destino.fill = copy(origem.fill)
        destino.border = copy(origem.border)
        destino.alignment = copy(origem.alignment)
        destino.number_format = copy(origem.number_format)
        destino.protection = copy(origem.protection)

def inserir_linha_formatada(ws, linha_modelo, linha_destino):

    for col in range(1, ws.max_column + 1):

        origem = ws.cell(
            row=linha_modelo,
            column=col
        )

        destino = ws.cell(
            row=linha_destino,
            column=col
        )

        copiar_estilo(origem, destino)

        destino.number_format = origem.number_format

    ws.row_dimensions[linha_destino].height = (
        ws.row_dimensions[linha_modelo].height
    )

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def style_range(ws, cell_range, border=None, fill=None, align=None):
    for row in ws[cell_range]:
        for cell in row:
            if border:
                cell.border = border
            if fill:
                cell.fill = fill
            if align:
                cell.alignment = align

def cm_para_px(cm):
    return int((cm / 2.54) * 96)

def normalizar_colunas(df):
    df.columns = [normalizar_texto(col) for col in df.columns]
    return df

def formatar_data(valor):
    if pd.isna(valor):
        return ""
    try:
        return pd.to_datetime(valor).strftime("%d/%m/%Y")
    except:
        return ""

def formatar_mes_pagto(data):

    if pd.isna(data):
        return ""

    meses = {
        1: "jan",
        2: "fev",
        3: "mar",
        4: "abr",
        5: "mai",
        6: "jun",
        7: "jul",
        8: "ago",
        9: "set",
        10: "out",
        11: "nov",
        12: "dez"
    }

    mes = meses[data.month]

    return f"{mes}/{data.year}"

def converter_valor_brasileiro(valor):

    if pd.isna(valor):
        return None

    # já numérico
    if isinstance(valor, (int, float)):
        return float(valor)

    valor = str(valor).strip()

    if valor == "":
        return None

    # remove espaços
    valor = valor.replace(" ", "")

    # CASO 1:
    # formato brasileiro
    # 1.234,56
    if "," in valor and "." in valor:
        valor = valor.replace(".", "")
        valor = valor.replace(",", ".")

    # CASO 2:
    # formato brasileiro simples
    # 123,45
    elif "," in valor:
        valor = valor.replace(",", ".")

    # CASO 3:
    # formato americano já correto
    # 123.45
    # não mexe

    try:
        return float(valor)

    except:
        return None

def selecionar_base():
    global base_file
    base_file = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx")]
    )
    if base_file:
        label_base.config(text=os.path.basename(base_file))
    atualizar_status()

def selecionar_extrato():
    global extrato_file
    extrato_file = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if extrato_file:
        label_extrato.config(text=os.path.basename(extrato_file))
    atualizar_status()

def selecionar_reports():
    global report_files
    report_files = filedialog.askopenfilenames(
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if report_files:
        label_reports.config(text=f"{len(report_files)} arquivos")
    atualizar_status()

def remover_base():
    global base_file
    base_file = None
    label_base.config(text="-")

def remover_extrato():
    global extrato_file
    extrato_file = None
    label_extrato.config(text="-")
    atualizar_status()

def remover_reports():
    global report_files
    report_files = []
    label_reports.config(text="-")
    atualizar_status()

def atualizar_status():
    btn_gerar.config(state="normal" if base_file or (extrato_file and report_files) else "disabled")

# ==============================
# PROCESSAR INFORMAÇÕES
# ==============================

def gerar_planilha():
    global base_file

    if base_file and not extrato_file and not report_files:
        wb = load_workbook(base_file)
        ws = wb[wb.sheetnames[0]]

        linha_header = validar_planilha_base(ws)

        if "Folha de Rosto" not in wb.sheetnames:
            criar_folha_de_rosto(wb, ws, linha_header)

            messagebox.showinfo(
                "Pronto",
                "Folha de rosto adicionada!"
            )

            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx"
            )

            if not path:
                return

            wb.save(path)

            if abrir_var.get():
                try:
                    os.startfile(path)

                except:
                    subprocess.call(["open", path])

            return

        else:

            messagebox.showinfo(
                "",
                "Folha de rosto já existe."
            )

    else:
        try:
            titulo_pc = entry_pc.get()
            head_line_extrato = detectar_header_extrato(extrato_file)

            df_extrato = ler_excel(extrato_file, header=head_line_extrato)
            df_extrato = normalizar_colunas(df_extrato)

            col_doc = encontrar_coluna(df_extrato, [
                "Numero Documento",
                "DOC"
            ])
            col_data = encontrar_coluna(df_extrato, [
                "Data",
                "DATA",
                "data"
            ])
            col_valor = encontrar_coluna(df_extrato, [
                "Valor R$",
                "Valor"
            ])
            col_inf = encontrar_coluna(df_extrato, [
                "Inf."
            ], obrigatoria=False)
            col_hist = encontrar_coluna(df_extrato, [
                "Histórico",
                "Historico",
                "Descrição",
                "Descricao"
            ], obrigatoria=False)

            if col_hist:
                df_extrato[col_hist] = df_extrato[col_hist].astype(str)

                df_extrato = df_extrato[
                    ~df_extrato[col_hist].str.lower().str.contains("saldo anterior|s a l d o", na=False)
                ]

            col_valor_original = f"{col_valor}_original"
            df_extrato[col_valor_original] = (
                df_extrato[col_valor].astype(str)
            )
            df_extrato[col_valor] = df_extrato[col_valor].apply(
                converter_valor_brasileiro
            )

            df_extrato[col_data] = pd.to_datetime(df_extrato[col_data], dayfirst=True, errors='coerce')

            if col_inf:
                df_extrato["valor_ajustado"] = df_extrato.apply(
                    lambda x: -abs(x[col_valor]) if str(x[col_inf]).strip() == "C"
                    else abs(x[col_valor]), axis=1
                )
            else:
                def ajustar_valor_tipo_b(row):
                    valor = abs(row[col_valor])

                    texto_original = str(
                        row.get(col_valor_original, "")
                    ).strip()

                    if "+" in texto_original or texto_original.startswith("+"):
                        return -valor

                    return valor

                df_extrato["valor_ajustado"] = df_extrato.apply(
                    ajustar_valor_tipo_b,
                    axis=1
                )

            dfs = []
            for f in report_files:
                df = ler_excel(f, header=5)
                df = normalizar_colunas(df)
                dfs.append(df)

            df_reports = pd.concat(dfs, ignore_index=True)

            col_data_rep = normalizar_texto("Vencimento/Mov.")
            col_valor_rep = normalizar_texto("Valor Total")
            col_tag = normalizar_texto("Rubrica")
            col_fornecedor = normalizar_texto("Razão Social")
            col_cnpj = normalizar_texto("CPF/CNPJ")
            col_emissao = normalizar_texto("Emissão")
            col_nf = normalizar_texto("Nº Documento")

            # col_fantasia = normalizar_texto("Fantasia")
            # col_operacao = normalizar_texto("Operação")
            col_pedido = normalizar_texto("Cód. Pedido")
            col_controle = normalizar_texto("Cód. Controle")
            col_hist_rep = normalizar_texto("Histórico")

            df_reports[col_valor_rep] = pd.to_numeric(df_reports[col_valor_rep], errors='coerce')
            df_reports[col_data_rep] = pd.to_datetime(df_reports[col_data_rep], dayfirst=True, errors='coerce')
            df_reports[col_emissao] = pd.to_datetime(df_reports[col_emissao], dayfirst=True, errors='coerce')

            resultado = []
            usados = set()

            for _, row in df_extrato.iterrows():
                valor = row["valor_ajustado"]
                data = row[col_data]

                if pd.isna(data) or pd.isna(valor):
                    continue  # pula essa linha do extrato

                candidatos = df_reports[
                    (df_reports[col_data_rep].notna()) &
                    (df_reports[col_valor_rep].notna())
                ].copy()

                candidatos = candidatos[
                    (abs(abs(candidatos[col_valor_rep]) - abs(valor)) < 0.01)
                ]

                # calcula diferença de datas com segurança
                candidatos["diff_dias"] = (candidatos[col_data_rep] - data).dt.days

                candidatos = candidatos[
                    candidatos["diff_dias"].notna() &
                    (abs(candidatos["diff_dias"]) <= 1)
                ]

                candidatos = candidatos[~candidatos.index.isin(usados)]

                excecao = ""

                if len(candidatos) == 1:
                    m = candidatos.iloc[0]
                    usados.add(m.name)

                elif len(candidatos) > 1:
                    if col_hist:
                        hist_ext = normalizar_texto(row.get(col_hist, ""))

                        candidatos["match_hist"] = candidatos.apply(
                            lambda x: hist_ext in normalizar_texto(x.get(col_hist_rep, "")),
                            axis=1
                        )

                        candidatos_match = candidatos[candidatos["match_hist"] == True]
                    else:
                        candidatos_match = pd.DataFrame()

                    if len(candidatos_match) == 1:
                        m = candidatos_match.iloc[0]
                        usados.add(m.name)
                        excecao = "Desempate por histórico"

                    else:
                        m = candidatos.iloc[0]
                        usados.add(m.name)
                        excecao = "Múltiplos matches (não resolvido)"

                else:
                    m = None

                if m is not None:
                    rubrica = "" if pd.isna(m[col_tag]) else str(m[col_tag]).title()
                    fornecedor = m.get(col_fornecedor, "")
                    cnpj = "" if pd.isna(m.get(col_cnpj, "")) else str(m.get(col_cnpj, ""))
                    data_nf = m.get(col_emissao, "")
                    nf = m.get(col_nf, "")
                    cod_pedido = m.get(col_pedido, "")
                    cod_controle = m.get(col_controle, "")
                    hist_rep = m.get(col_hist_rep, "")
                else:
                    rubrica = fornecedor = cnpj = data_nf = nf = cod_pedido = cod_controle = hist_rep = ""
                    excecao = "Sem match"

                doc_num = pd.to_numeric(row[col_doc], errors='coerce')

                resultado.append({
                    "Nº": len(resultado) + 1,
                    "RESUMO DO GASTO": "-",
                    "RUBRICA": rubrica,
                    "FORNECEDOR": fornecedor,
                    "PC": titulo_pc,
                    "CNPJ ou CPF": cnpj,
                    "DATA DA NF": formatar_data(data_nf),
                    "(número) NF/RECIBO": nf,
                    "Nº EXTRATO": doc_num,
                    "DATA DO PAGAMENTO": formatar_data(data),
                    "VALOR PAGO": valor,
                    "CÓD. PEDIDO": cod_pedido,
                    "CÓD. RMS": cod_controle,
                    "Mês_pagto": formatar_mes_pagto(data),
                    "HISTÓRICO": hist_rep,
                    "EXCEÇÕES DO ALGORITMO": excecao
                })

            df_final = pd.DataFrame(resultado)
            df_final = df_final.replace({pd.NA: "", "nan": "", "NaN": ""})

            messagebox.showinfo("Pronto", "Planilha criada!")

            salvar_arquivo(df_final)

        except Exception as e:
            messagebox.showerror("Erro", str(e))

# ==============================
# SALVAR PLANILHA
# ==============================

def salvar_arquivo(df):
    global base_file

    path = filedialog.asksaveasfilename(
        defaultextension=".xlsx"
    )

    if not path:
        return

    if base_file:

        wb = load_workbook(base_file)
        ws = wb[wb.sheetnames[0]]

        linha_header = validar_planilha_base(ws)

        info = obter_info_tabela(ws, linha_header)
        mapa = info["mapa"]

        ultima_linha_base = detectar_ultima_linha(
            ws,
            mapa,
            linha_header + 1
        )

        linha_insercao = ultima_linha_base + 1

        anotacoes = []

        if linha_insercao <= ws.max_row:

            for row in range(
                linha_insercao,
                ws.max_row + 1
            ):

                linha = []

                for col in range(
                    1,
                    ws.max_column + 1
                ):

                    cell = ws.cell(
                        row=row,
                        column=col
                    )

                    linha.append({
                        "valor": cell.value,
                        "font": copy(cell.font),
                        "fill": copy(cell.fill),
                        "border": copy(cell.border),
                        "alignment": copy(cell.alignment),
                        "number_format": cell.number_format,
                        "protection": copy(cell.protection)
                    })

                anotacoes.append({
                    "altura": ws.row_dimensions[row].height,
                    "dados": linha
                })

            ws.delete_rows(
                linha_insercao,
                ws.max_row - linha_insercao + 1
            )

        linha_modelo = linha_header + 1

        ultimo_indice = 0

        col_indice = mapa["no"][0]

        for row in range(
            linha_header + 1,
            ultima_linha_base + 1
        ):

            valor = ws.cell(
                row=row,
                column=col_indice
            ).value

            try:
                valor = int(valor)

                if valor > ultimo_indice:
                    ultimo_indice = valor

            except:
                pass

        for _, row_df in df.iterrows():

            inserir_linha_formatada(ws, linha_modelo, linha_insercao)

            for col in range(1, ws.max_column + 1):
                ws.cell(
                    row=linha_insercao,
                    column=col
                ).fill = PatternFill(fill_type=None)

            for nome_coluna, colunas_ws in mapa.items():

                if nome_coluna in associacoes:

                    coluna_df = associacoes[nome_coluna]

                    if coluna_df in row_df.index:

                        valor = row_df[coluna_df]

                        if pd.isna(valor):
                            valor = ""

                        for coluna_ws in colunas_ws:

                            ws.cell(
                                row=linha_insercao,
                                column=coluna_ws,
                                value=valor
                            )

            ultimo_indice += 1

            ws.cell(
                row=linha_insercao,
                column=col_indice,
                value=ultimo_indice
            )

            linha_insercao += 1

        for linha in anotacoes:

            ws.row_dimensions[
                linha_insercao
            ].height = linha["altura"]

            for col_idx, dados in enumerate(
                linha["dados"],
                1
            ):

                cell = ws.cell(
                    row=linha_insercao,
                    column=col_idx,
                    value=dados["valor"]
                )

                cell.font = copy(dados["font"])
                cell.fill = copy(dados["fill"])
                cell.border = copy(dados["border"])
                cell.alignment = copy(dados["alignment"])
                cell.number_format = dados["number_format"]
                cell.protection = copy(dados["protection"])

            linha_insercao += 1

        col_valor = mapa["valor pago"][0]

        letra_valor = get_column_letter(
            col_valor
        )

        primeira_linha_dados = linha_header + 1

        ultima_linha_tabela = (
            linha_insercao
            - len(anotacoes)
            - 1
        )

        ws[f"{letra_valor}1"] = (
            f"=SUM("
            f"{letra_valor}{primeira_linha_dados}:"
            f"{letra_valor}{ultima_linha_tabela}"
            f")"
        )

        ws[f"{letra_valor}1"].number_format = (
            '#,##0.00'
        )

        if "Folha de Rosto" in wb.sheetnames:
            del wb["Folha de Rosto"]
        
        criar_folha_de_rosto(wb, ws, linha_header)

        wb.save(path)

    else:

        wb = Workbook()

        ws = wb.active

        ws.title = entry_pc.get()

        ws.sheet_view.showGridLines = False

        for c, col in enumerate(
            df.columns,
            1
        ):

            ws.cell(
                row=2,
                column=c,
                value=col
            )

        for r, row in enumerate(
            df.values,
            3
        ):

            for c, val in enumerate(
                row,
                1
            ):

                if str(val).lower() == "nan":
                    val = ""

                ws.cell(
                    row=r,
                    column=c,
                    value=val
                )

        larguras = {
            'A': 7.11+0.78,
            'B': 50.33+0.78,
            'C': 32.56+0.78,
            'D': 51.56+0.78,
            'E': 6.22+0.78,
            'F': 28.11+0.78,
            'G': 16.89+0.78,
            'H': 27.56+0.78,
            'I': 26.67+0.78,
            'J': 29.67+0.78,
            'K': 14.67+0.78,
            'L': 19.89+0.78,
            'M': 11.89+0.78,
            'N': 16.11+0.78,
            'O': 16.67+0.78,
            'P': 36.67+0.78
        }

        for col, largura in larguras.items():

            ws.column_dimensions[
                col
            ].width = largura

        ws.row_dimensions[2].height = 29.4

        for r in range(3, ws.max_row + 1):

            ws.row_dimensions[r].height = 14.4

        for row in ws.iter_rows(
            min_row=2,
            max_row=ws.max_row
        ):

            for cell in row:

                cell.border = border

                cell.alignment = Alignment(
                    wrap_text=True
                )

        for r in range(3, ws.max_row + 1):

            ws.cell(
                row=r,
                column=9
            ).number_format = '#,##0'

            ws.cell(
                row=r,
                column=11
            ).number_format = '#,##0.00'

        ultima_linha = ws.max_row

        ws["K1"] = (
            f"=SUM(K3:K{ultima_linha})"
        )

        ws["K1"].number_format = '#,##0.00'

        ws["K1"].alignment = Alignment(
            horizontal="right"
        )

        for c in range(1, ws.max_column + 1):
            ws.cell(row=2, column=c).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=2, column=c).font = Font(bold=True)

        for row in range(3, ws.max_row + 1):

            for col, alinhamento in alinhamentos.items():

                ws[f"{col}{row}"].alignment = Alignment(
                    horizontal=alinhamento,
                    vertical="center",
                    wrap_text=True
                )

        criar_folha_de_rosto(wb, ws, 2)

        wb.save(path)

    if abrir_var.get():
        try:
            os.startfile(path)

        except:
            subprocess.call(["open", path])

# ==============================
# INTERFACE (UI)
# ==============================

# ==============================
#   JANELA

root = ttk.Window(themename="darkly")
root.title("HBR PC Assist")
root.iconbitmap(resource_path("hbr_pc_assist.ico"))
root.geometry("720x520")
root.resizable(True, True)

# ==============================
#    ESTILOS CUSTOM

style = ttk.Style()

style.configure(
    "Accent.TButton",
    background="#f9b02e",
    foreground="black",
    borderwidth=0,
    focusthickness=3,
    focuscolor="none",
    padding=8
)

style.map(
    "Accent.TButton",
    background=[
        ("active", "#ffd27a"),   # hover
        ("pressed", "#d9941f")   # clique
    ]
)

style.configure(
    "Title.TLabel",
    font=("Segoe UI", 16, "bold")
)

style.configure(
    "Subtitle.TLabel",
    font=("Segoe UI", 11, "bold")
)

style.configure(
    "Status.TLabel",
    font=("Segoe UI", 9)
)

# ==============================
#    LAYOUT PRINCIPAL

root.rowconfigure(0, weight=1)
root.columnconfigure(0, weight=1)

wrapper = ttk.Frame(root)
wrapper.grid(row=0, column=0, sticky="nsew")

wrapper.rowconfigure(0, weight=1)
wrapper.columnconfigure(0, weight=1)

container = ttk.Frame(wrapper, padding=20)
container.grid(row=0, column=0)

# GRID config
for i in range(4):
    container.columnconfigure(i, weight=1)

# ==============================
#    TÍTULO

ttk.Label(
    container,
    text="HBR PC Assist",
    style="Title.TLabel"
).grid(row=0, column=0, columnspan=4, pady=(0, 20))

# ==============================
#    BLOCO: BASE

ttk.Label(container, text="Planilha Base (opcional)", style="Subtitle.TLabel")\
    .grid(row=1, column=0, sticky="w")

ttk.Button(container, text="Selecionar", style="Accent.TButton",
           command=selecionar_base)\
    .grid(row=1, column=1, padx=5)

ttk.Button(container, text="Remover", style="Accent.TButton",
           command=remover_base)\
    .grid(row=1, column=2, padx=5)

label_base = ttk.Label(container, text="-", style="Status.TLabel")
label_base.grid(row=2, column=0, columnspan=4, sticky="w", pady=(0, 10))

# ==============================
#    BLOCO: EXTRATO

ttk.Label(container, text="Extrato CC", style="Subtitle.TLabel")\
    .grid(row=3, column=0, sticky="w")

ttk.Button(container, text="Selecionar", style="Accent.TButton",
           command=selecionar_extrato)\
    .grid(row=3, column=1, padx=5)

ttk.Button(container, text="Remover", style="Accent.TButton",
           command=remover_extrato)\
    .grid(row=3, column=2, padx=5)

label_extrato = ttk.Label(container, text="-", style="Status.TLabel")
label_extrato.grid(row=4, column=0, columnspan=4, sticky="w", pady=(0, 10))

# ==============================
#    BLOCO: REPORTS

ttk.Label(container, text="Relatórios", style="Subtitle.TLabel")\
    .grid(row=5, column=0, sticky="w")

ttk.Button(container, text="Selecionar", style="Accent.TButton",
           command=selecionar_reports)\
    .grid(row=5, column=1, padx=5)

ttk.Button(container, text="Remover", style="Accent.TButton",
           command=remover_reports)\
    .grid(row=5, column=2, padx=5)

label_reports = ttk.Label(container, text="-", style="Status.TLabel")
label_reports.grid(row=6, column=0, columnspan=4, sticky="w", pady=(0, 20))

# ==============================
#    INPUT PC

ttk.Label(container, text="Título PC", style="Subtitle.TLabel")\
    .grid(row=7, column=0, sticky="w")

entry_pc = ttk.Entry(container)
entry_pc.grid(row=8, column=0, columnspan=4, sticky="ew", pady=(0, 15))

# ==============================
#    CHECKBOX

abrir_var = ttk.BooleanVar()

ttk.Checkbutton(
    container,
    text="Abrir planilha quando estiver pronta",
    variable=abrir_var,
    bootstyle="round-toggle"
).grid(row=9, column=0, columnspan=4, sticky="w", pady=(0, 15))

# ==============================
#   BOTÃO FINAL

btn_gerar = ttk.Button(
    container,
    text="Gerar Planilha",
    style="Accent.TButton",
    state=DISABLED,
    command=gerar_planilha
)

btn_gerar.grid(row=11, column=0, columnspan=4, pady=10)

# ==============================
#    RUN

root.mainloop()
